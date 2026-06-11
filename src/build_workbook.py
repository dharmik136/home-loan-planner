"""Generate the live Home Loan Prepayment Planner workbook.

Fully formula-driven: the user edits the Inputs and the per-month Prepay columns
inside Excel and every schedule, summary, and windfall figure recalculates.
The Python amortization core (amortization.py) is the reference used to VERIFY
these formulas (see verify_workbook.py) — it is not what fills the cells.
"""
from __future__ import annotations

import datetime as dt
import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

N = 180  # max months (15-year tenure)
DATA0 = 3  # first data row in schedule sheets (row 1 title, row 2 header)

# ---- styling helpers ---------------------------------------------------------
TITLE = Font(bold=True, size=14, color="1F3864")
HEAD = Font(bold=True, color="FFFFFF")
HEAD_FILL = PatternFill("solid", fgColor="1F3864")
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")   # yellow = editable
SAVE_FILL = PatternFill("solid", fgColor="E2EFDA")    # green = headline result
LABEL = Font(bold=True)
RUPEE = '#,##0'
RUPEE2 = '#,##0.00'
thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def style_header(ws, row, first_col, last_col):
    for c in range(first_col, last_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEAD
        cell.fill = HEAD_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = BORDER


# ---- Inputs sheet ------------------------------------------------------------
def build_inputs(ws):
    ws["A1"] = "INPUTS — edit the yellow cells"
    ws["A1"].font = TITLE
    ws.cell(row=2, column=2, value="Loan A").font = LABEL
    ws.cell(row=2, column=3, value="Loan B").font = LABEL
    rows = [
        ("Outstanding principal (Rs)", 3_000_000, 5_000_000, RUPEE, True),
        ("Annual interest rate (%)", 7.5, 7.5, "0.00", True),
        ("Tenure (months)", 180, 180, "0", True),
        ("EMI start date", dt.date(2026, 6, 1), dt.date(2026, 6, 1), "dd-mmm-yyyy", True),
    ]
    r = 3
    for label, a, b, fmt, editable in rows:
        ws.cell(row=r, column=1, value=label).font = LABEL
        for col, val in ((2, a), (3, b)):
            cell = ws.cell(row=r, column=col, value=val)
            cell.number_format = fmt
            if editable:
                cell.fill = INPUT_FILL
            cell.border = BORDER
        r += 1
    # Auto rows: EMI (rounded up) and monthly rate
    for col in (2, 3):
        L = get_column_letter(col)
        ws.cell(row=7, column=1, value="Monthly EMI (auto)").font = LABEL
        ws.cell(row=7, column=col,
                value=f"=ROUNDUP({L}3*({L}4/100/12)*POWER(1+{L}4/100/12,{L}5)"
                      f"/(POWER(1+{L}4/100/12,{L}5)-1),0)").number_format = RUPEE
        ws.cell(row=8, column=1, value="Monthly rate (auto)").font = LABEL
        ws.cell(row=8, column=col, value=f"={L}4/100/12").number_format = "0.000000"
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 16


# ---- Schedule sheet (Baseline + Plan side by side) ---------------------------
SCHED_HEADERS = [
    "Month", "Date",
    "Opening", "Interest", "EMI", "Principal", "Closing",          # baseline C-G
    "PREPAY (edit)", "Rule check",                                  # plan input H-I
    "Opening", "Interest", "EMI", "Principal", "Prepay", "Closing",  # plan J-P
]


def build_schedule(ws, inp_col):
    """inp_col: 'B' for Loan A, 'C' for Loan B (column on Inputs sheet)."""
    P = f"Inputs!${inp_col}$3"
    EMI = f"Inputs!${inp_col}$7"
    MR = f"Inputs!${inp_col}$8"
    START = f"Inputs!${inp_col}$6"

    ws["A1"] = f"AMORTIZATION SCHEDULE — Loan {'A' if inp_col=='B' else 'B'}"
    ws["A1"].font = TITLE
    ws.cell(row=2, column=1, value="").value = None
    # Two group headers
    ws.cell(row=1, column=4, value="(C-G: baseline, no prepayment   |   H: type any extra payment   |   J-P: your plan)")
    ws.cell(row=1, column=4).font = Font(italic=True, size=9, color="808080")
    for i, h in enumerate(SCHED_HEADERS, start=1):
        ws.cell(row=2, column=i, value=h)
    style_header(ws, 2, 1, len(SCHED_HEADERS))

    for m in range(1, N + 1):
        r = DATA0 + m - 1
        prev = r - 1
        # A month, B date
        ws.cell(row=r, column=1, value=m).border = BORDER
        ws.cell(row=r, column=2, value=f"=EDATE({START},A{r}-1)").number_format = "dd-mmm-yyyy"
        # --- baseline C-G ---
        opening_b = f"={P}" if m == 1 else f"=G{prev}"
        ws.cell(row=r, column=3, value=opening_b)
        ws.cell(row=r, column=4, value=f"=IF(C{r}<=0.005,0,ROUND(C{r}*{MR},2))")
        ws.cell(row=r, column=5, value=f"=IF(C{r}<=0.005,0,MIN({EMI},C{r}+D{r}))")
        ws.cell(row=r, column=6, value=f"=E{r}-D{r}")
        ws.cell(row=r, column=7, value=f"=ROUND(C{r}-F{r},2)")
        # --- plan input H-I ---
        h = ws.cell(row=r, column=8, value=0)
        h.fill = INPUT_FILL
        ws.cell(row=r, column=9,
                value=(f'=IF(H{r}=0,"",IF(A{r}<2,"X not before month 2",'
                       f'IF(H{r}<MAX(5000,{EMI}),"X below min",'
                       f'IF(H{r}>MIN(5000000,0.75*J{r}),"X above max","OK"))))'))
        # --- plan J-P ---
        opening_p = f"={P}" if m == 1 else f"=O{prev}"  # plan closing is column O
        ws.cell(row=r, column=10, value=opening_p)
        ws.cell(row=r, column=11, value=f"=IF(J{r}<=0.005,0,ROUND(J{r}*{MR},2))")
        ws.cell(row=r, column=12, value=f"=IF(J{r}<=0.005,0,MIN({EMI},J{r}+K{r}))")
        ws.cell(row=r, column=13, value=f"=L{r}-K{r}")
        ws.cell(row=r, column=14, value=f"=IF(J{r}<=0.005,0,MIN(H{r},J{r}-M{r}))")
        ws.cell(row=r, column=15, value=f"=ROUND(J{r}-M{r}-N{r},2)")
        for c in range(3, 16):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER
            if c != 9:
                cell.number_format = RUPEE2
    widths = {"A": 6, "B": 12}
    for col in range(3, 16):
        widths[get_column_letter(col)] = 12
    for k, v in widths.items():
        ws.column_dimensions[k].width = v
    ws.freeze_panes = "C3"


# ---- Summary sheet -----------------------------------------------------------
def build_summary(ws):
    ws["A1"] = "SUMMARY — baseline vs your prepayment plan"
    ws["A1"].font = TITLE
    last = DATA0 + N - 1

    ws.cell(row=3, column=2, value="Loan A").font = LABEL
    ws.cell(row=3, column=3, value="Loan B").font = LABEL
    ws.cell(row=3, column=4, value="Combined").font = LABEL

    def refs(sheet):
        return (
            f"SUM({sheet}!D{DATA0}:D{last})",                  # baseline interest
            f"SUM({sheet}!K{DATA0}:K{last})",                  # plan interest
            f'COUNTIF({sheet}!E{DATA0}:E{last},">0.005")',     # baseline payoff months
            f'COUNTIF({sheet}!L{DATA0}:L{last},">0.005")',     # plan payoff months
        )

    a = refs("Schedule_A")
    b = refs("Schedule_B")

    # (row, label, Loan A, Loan B, Combined, number_format, highlight)
    spec = [
        (4, "Outstanding (Rs)", "=Inputs!B3", "=Inputs!C3", "=B4+C4", RUPEE, False),
        (5, "Monthly EMI (Rs)", "=Inputs!B7", "=Inputs!C7", "=B5+C5", RUPEE, False),
        (6, "Baseline total interest (Rs)", f"={a[0]}", f"={b[0]}", "=B6+C6", RUPEE, False),
        (7, "Plan total interest (Rs)", f"={a[1]}", f"={b[1]}", "=B7+C7", RUPEE, False),
        (8, "INTEREST SAVED (Rs)", "=B6-B7", "=C6-C7", "=B8+C8", RUPEE, True),
        (9, "Baseline payoff (months)", f"={a[2]}", f"={b[2]}", None, "0", False),
        (10, "Plan payoff (months)", f"={a[3]}", f"={b[3]}", None, "0", False),
        (11, "MONTHS SAVED", "=B9-B10", "=C9-C10", None, "0", True),
        (12, "Years saved", "=B11/12", "=C11/12", None, "0.0", True),
        (13, "Baseline payoff date", "=EDATE(Inputs!B6,B9-1)", "=EDATE(Inputs!C6,C9-1)", None, "dd-mmm-yyyy", False),
        (14, "Plan payoff date", "=EDATE(Inputs!B6,B10-1)", "=EDATE(Inputs!C6,C10-1)", None, "dd-mmm-yyyy", False),
    ]
    for row, lab, av, bv, cv, fmt, hl in spec:
        c = ws.cell(row=row, column=1, value=lab)
        c.font = LABEL
        if hl:
            c.fill = SAVE_FILL
        for col, val in ((2, av), (3, bv), (4, cv)):
            if val is None:
                continue
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = BORDER
            cell.number_format = fmt
            if hl:
                cell.fill = SAVE_FILL

    # Recommendation: higher rate first; tie -> larger outstanding.
    ws.cell(row=16, column=1, value="Prepay FIRST:").font = LABEL
    ws.cell(row=16, column=2,
            value='=IF(Inputs!B4>Inputs!C4,"Loan A (higher rate)",'
                  'IF(Inputs!C4>Inputs!B4,"Loan B (higher rate)",'
                  'IF(Inputs!B3>=Inputs!C3,"Loan B (larger balance saves more)",'
                  '"Loan A (larger balance saves more)")))')
    ws.cell(row=16, column=2).fill = SAVE_FILL
    ws.merge_cells("B16:D16")
    ws.column_dimensions["A"].width = 30
    for c in "BCD":
        ws.column_dimensions[c].width = 16


# ---- Windfall What-If sheet --------------------------------------------------
def build_windfall(ws):
    ws["A1"] = "WINDFALL WHAT-IF — one lump sum, which loan saves more?"
    ws["A1"].font = TITLE
    ws.cell(row=3, column=1, value="Lump sum (Rs):").font = LABEL
    c = ws.cell(row=3, column=2, value=500_000); c.fill = INPUT_FILL; c.number_format = RUPEE
    ws.cell(row=4, column=1, value="Apply in month #:").font = LABEL
    c = ws.cell(row=4, column=2, value=12); c.fill = INPUT_FILL
    last = DATA0 + N - 1

    hdr = DATA0 + N + 2          # block header row, below the result area
    d0 = hdr + 1                 # block data start
    dN = d0 + N - 1

    # Two compute blocks: A in cols A-F, B in cols H-M. Col A/H = month.
    def block(start_col, inp_col):
        L0 = get_column_letter(start_col)       # Month
        Lo = get_column_letter(start_col + 1)   # Opening
        Li = get_column_letter(start_col + 2)   # Interest
        Le = get_column_letter(start_col + 3)   # EMI
        Lp = get_column_letter(start_col + 4)   # PrepayApplied
        Lc = get_column_letter(start_col + 5)   # Closing
        P = f"Inputs!${inp_col}$3"; EMI = f"Inputs!${inp_col}$7"; MR = f"Inputs!${inp_col}$8"
        for h, lab in zip(
            [L0, Lo, Li, Le, Lp, Lc],
            ["Month", "Opening", "Interest", "EMI", "Prepay", "Closing"],
        ):
            ws[f"{h}{hdr}"] = lab
        style_header(ws, hdr, start_col, start_col + 5)
        for m in range(1, N + 1):
            r = d0 + m - 1
            ws[f"{L0}{r}"] = m
            ws[f"{Lo}{r}"] = f"={P}" if m == 1 else f"={Lc}{r-1}"
            ws[f"{Li}{r}"] = f"=IF({Lo}{r}<=0.005,0,ROUND({Lo}{r}*{MR},2))"
            ws[f"{Le}{r}"] = f"=IF({Lo}{r}<=0.005,0,MIN({EMI},{Lo}{r}+{Li}{r}))"
            # prepay only in the chosen month, capped at remaining principal
            ws[f"{Lp}{r}"] = (f"=IF({Lo}{r}<=0.005,0,IF({L0}{r}=$B$4,"
                              f"MIN($B$3,{Lo}{r}-({Le}{r}-{Li}{r})),0))")
            ws[f"{Lc}{r}"] = f"=ROUND({Lo}{r}-({Le}{r}-{Li}{r})-{Lp}{r},2)"
            for cc in range(start_col + 1, start_col + 6):
                ws.cell(row=r, column=cc).number_format = RUPEE2
        return Li, Le, dN  # interest col, emi col

    aint, aemi, _ = block(1, "B")   # Loan A block cols A-F
    bint, bemi, _ = block(8, "C")   # Loan B block cols H-M

    # Results referencing baseline from Schedule sheets vs these windfall blocks.
    base_int_a = f"SUM(Schedule_A!D{DATA0}:D{last})"
    base_int_b = f"SUM(Schedule_B!D{DATA0}:D{last})"
    base_m_a = f'COUNTIF(Schedule_A!E{DATA0}:E{last},">0.005")'
    base_m_b = f'COUNTIF(Schedule_B!E{DATA0}:E{last},">0.005")'
    wf_int_a = f"SUM({aint}{d0}:{aint}{dN})"
    wf_int_b = f"SUM({bint}{d0}:{bint}{dN})"
    wf_m_a = f'COUNTIF({aemi}{d0}:{aemi}{dN},">0.005")'
    wf_m_b = f'COUNTIF({bemi}{d0}:{bemi}{dN},">0.005")'

    res = [
        (6, "Loan A — interest saved (Rs):", f"={base_int_a}-{wf_int_a}", RUPEE),
        (7, "Loan A — months saved:", f"={base_m_a}-{wf_m_a}", "0"),
        (8, "Loan B — interest saved (Rs):", f"={base_int_b}-{wf_int_b}", RUPEE),
        (9, "Loan B — months saved:", f"={base_m_b}-{wf_m_b}", "0"),
    ]
    for r, lab, formula, fmt in res:
        ws.cell(row=r, column=1, value=lab).font = LABEL
        cell = ws.cell(row=r, column=2, value=formula)
        cell.number_format = fmt
        cell.border = BORDER
    ws.cell(row=10, column=1, value="RECOMMENDATION:").font = LABEL
    rec = ws.cell(row=10, column=2,
                  value='=IF(B6>B8,"Put it on Loan A (saves more interest)",'
                        'IF(B8>B6,"Put it on Loan B (saves more interest)","Either — equal saving"))')
    rec.fill = SAVE_FILL
    ws.merge_cells("B10:F10")
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["G"].width = 3


# ---- HDFC Rules + README -----------------------------------------------------
def build_rules(ws):
    ws["A1"] = "HDFC HOME-LOAN PREPAYMENT RULES (verified June 2026)"
    ws["A1"].font = TITLE
    data = [
        ("Rule", "Value / note"),
        ("Prepayment charge (floating)", "NONE. RBI ban from 1 Jan 2026 on floating loans to individuals (non-business) — any amount, any source."),
        ("Frequency", "Once per calendar month (up to 12x / year)."),
        ("Minimum part payment", "Rs 5,000 or 1 EMI, whichever is HIGHER."),
        ("Maximum part payment", "Rs 50 lakh/month OR 75% of year-opening principal, whichever is LOWER."),
        ("Timing", "Allowed only after 1 month of EMI commencement (not in month 1)."),
        ("Source of funds", "From the repayment-tagged bank account (or branch visit)."),
        ("Default after prepay", "Tenure reduction (EMI fixed, loan shortens) — used by this model."),
    ]
    for i, (a, b) in enumerate(data, start=3):
        ws.cell(row=i, column=1, value=a).font = LABEL if i > 3 else HEAD
        cb = ws.cell(row=i, column=2, value=b)
        cb.alignment = Alignment(wrap_text=True, vertical="top")
        if i == 3:
            style_header(ws, 3, 1, 2)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 70
    ws.cell(row=12, column=1,
            value="Sources: HDFC FAQ (homeloans.hdfc.bank.in/checklist/faqs), HDFC prepayment page, "
                  "RBI 2026 directive. Note: 'Max' check uses the prepay-month opening balance as a "
                  "proxy for year-opening principal.").font = Font(italic=True, size=9, color="808080")


def build_readme(ws):
    ws["A1"] = "HOME LOAN PREPAYMENT PLANNER"
    ws["A1"].font = TITLE
    lines = [
        "",
        "WHAT THIS DOES — answers: if I pay extra in the early years, how many years and",
        "how much interest do I save, and which of my two loans should I attack first?",
        "",
        "HOW TO USE:",
        "1. Inputs tab — set each loan's outstanding, rate, tenure, start date (yellow cells).",
        "   The EMI is computed automatically.",
        "2. Schedule_A / Schedule_B — to test a prepayment, type the extra amount in the",
        "   yellow PREPAY column next to the month you'd pay it. The 'Rule check' column",
        "   warns if it breaks an HDFC rule. The loan shortens automatically (tenure reduction).",
        "3. Summary tab — see baseline vs your plan: interest saved, months/years saved,",
        "   new payoff date, and which loan to prepay first.",
        "4. Windfall What-If — got a one-time lump sum? Enter the amount + month and it tells",
        "   you which loan that lump saves the most on.",
        "5. HDFC Rules — the real rules this model enforces.",
        "",
        "KEY FACT: prepaying EARLY saves far more than the same amount paid later, because",
        "early EMIs are mostly interest. Floating-rate loans have NO prepayment penalty (RBI, 2026).",
        "",
        "Yellow = you edit.  Green = headline result.  All numbers recalculate live in Excel.",
    ]
    for i, t in enumerate(lines, start=2):
        ws.cell(row=i, column=1, value=t)
    ws.column_dimensions["A"].width = 95


def build(path: str):
    wb = Workbook()
    wb.remove(wb.active)
    build_readme(wb.create_sheet("README"))
    build_inputs(wb.create_sheet("Inputs"))
    build_schedule(wb.create_sheet("Schedule_A"), "B")
    build_schedule(wb.create_sheet("Schedule_B"), "C")
    build_summary(wb.create_sheet("Summary"))
    build_windfall(wb.create_sheet("Windfall What-If"))
    build_rules(wb.create_sheet("HDFC Rules"))
    wb.save(path)
    return path


if __name__ == "__main__":
    out = os.path.join(os.path.dirname(__file__), "..", "Home-Loan-Prepayment-Planner.xlsx")
    out = os.path.abspath(out)
    build(out)
    print("wrote", out)
